from llm_controller.llm_agent_action import *
from llm_controller.llm_agent_negotiation_system import *  # system perspective negotiation
from llm_controller.memory import DrivingMemory
import highway_env
import imageio
import openpyxl
import os
import shutil

def open_excel(i):
    file_dir = './llm_controller/excel/' + '/'
    file_name = file_dir + str(i) + '.xlsx'

    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    workbook = openpyxl.Workbook()
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)

    # if 'Sheet' in workbook.sheetnames:
    #     del workbook['Sheet']
    return file_name, workbook

def write_data(workbook, env, t):
    column_names = ['t', 'x', 'y', 'v', 'theta', 'background_veh?']
    for vehicle in env.road.vehicles:
        sheet_name = str(vehicle.id)
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(column_names)
        else:
            worksheet = workbook[sheet_name]
        controlled_vehicles = env.controlled_vehicles
        if vehicle not in controlled_vehicles:
            background_vehicles = False
        else:
            background_vehicles = True
        state = [round(vehicle.position[0], 2), round(vehicle.position[1], 2), round(vehicle.speed, 2), round(vehicle.heading, 2), background_vehicles]
        row_data = [t, round(vehicle.position[0], 2), round(vehicle.position[1], 2), round(vehicle.speed, 2), round(vehicle.heading, 2), background_vehicles]
        worksheet.append(row_data)
        worksheet.cell(row=t + 2, column=1, value=t)
        for i, item in enumerate(state):
            worksheet.cell(row=t + 2, column=i + 2, value=item)
    return workbook

# if choose merge, active config
# config = {
#     "simulation_frequency": 20,
#     "policy_frequency": 5,
#     "duration":40
# }

# env = gym.make('merge-multi-agent-v0', config=config)
env = gym.make('intersection-multi-agent-v0')
# env = gym.make('highway-v0')


for i in range(100):
    video_path = './llm_controller/video/' + str(i) + '.mp4'  
    writer = imageio.get_writer(video_path, fps=30) 
    file_name, workbook = open_excel(i)
    terminated = False
    t = 0
    obs = env.reset()
    while not (terminated):
        print('---------------------------------------------------------------')
        # memory module
        memory = DrivingMemory(env)

        # negotiation module
        llm_agent_conflict_resolver = LlmAgent_negotiation_module(env)  # system perspective negotiation
        negotiation_prompt, conflicting_info = llm_agent_conflict_resolver.llm_controller_run(env)

        # decision
        llm_agent = LlmAgent_action_module(env)  # create llmagent  for highway and merge env (can change lane)
        sce = llm_agent.retrun_sce()  # sce data
        llm_actions = llm_agent.llm_controller_run(env, negotiation_prompt, conflicting_info, env.controlled_vehicles, memory)  # negotiation results from upper layer and conflict info which stores distance speed


        action = [item for sublist in llm_actions for item in sublist]  # [[1], [3]]->[1, 3]

        for veh in env.controlled_vehicles:
            print(veh, veh.speed)

        obs, global_reward, terminated, info = env.step(tuple(action), env)
        env.render()
        print("llm_actions:", action)
        print("global_reward_llm:", global_reward)

        frame = env.render('rgb_array')
        writer.append_data(frame)

        workbook = write_data(workbook, env, t)
        workbook.save(file_name)
        t += 1

        # action = [1,3,1,3]  # env.action_space.sample()
        # obs, reward, terminated, info = env.step(tuple(action), env)
        # print(action)
        # print(terminated, info)
        env.render()
    writer.close()
    print(i)

